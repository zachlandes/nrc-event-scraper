"""Convert all data/events/*.jsonl files into a single Excel workbook."""

import json
import glob
import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Regex to strip illegal XML characters that openpyxl rejects
ILLEGAL_XML_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x84\x86-\x9f'
    r'\ud800-\udfff\ufdd0-\ufdef\ufffe\uffff]'
)


def sanitize(value):
    """Remove illegal XML characters from string values."""
    if isinstance(value, str):
        return ILLEGAL_XML_RE.sub('', value)
    return value

EVENTS_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "events")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "..", "data", "nrc_events_all.xlsx")


def flatten_event(event: dict) -> dict:
    """Flatten a single JSONL event record into a flat dict for tabular output."""
    row = {}

    # Scalar fields
    scalar_fields = [
        "event_number", "category", "report_date", "facility", "region", "state",
        "unit", "rx_type", "licensee", "license_number", "rep_org", "agreement",
        "docket", "county", "city",
        "notification_date", "notification_time", "notification_timezone",
        "event_date", "event_time", "event_timezone",
        "last_update_date", "emergency_class",
        "event_text", "scraped_at", "html_format",
    ]
    for field in scalar_fields:
        row[field] = event.get(field)

    # Extract year from event_date or report_date
    event_date = event.get("event_date") or event.get("report_date") or ""
    row["event_year"] = int(event_date[:4]) if len(event_date) >= 4 else None

    # Flatten cfr_sections: join as "code: description" entries
    cfr = event.get("cfr_sections", [])
    row["cfr_codes"] = "; ".join(s.get("code", "") for s in cfr) if cfr else None
    row["cfr_descriptions"] = "; ".join(s.get("description", "") for s in cfr) if cfr else None

    # Flatten persons_notified
    persons = event.get("persons_notified", [])
    row["persons_notified_names"] = "; ".join(p.get("name", "") for p in persons) if persons else None
    row["persons_notified_orgs"] = "; ".join(p.get("organization", "") for p in persons) if persons else None

    # Flatten reactor_units (up to 3 units)
    units = event.get("reactor_units", [])
    for i in range(3):
        prefix = f"reactor_unit_{i+1}_"
        if i < len(units):
            u = units[i]
            row[prefix + "unit"] = u.get("unit")
            row[prefix + "scram_code"] = u.get("scram_code")
            row[prefix + "rx_crit"] = u.get("rx_crit")
            row[prefix + "initial_power"] = u.get("initial_power")
            row[prefix + "initial_rx_mode"] = u.get("initial_rx_mode")
            row[prefix + "current_power"] = u.get("current_power")
            row[prefix + "current_rx_mode"] = u.get("current_rx_mode")
        else:
            row[prefix + "unit"] = None
            row[prefix + "scram_code"] = None
            row[prefix + "rx_crit"] = None
            row[prefix + "initial_power"] = None
            row[prefix + "initial_rx_mode"] = None
            row[prefix + "current_power"] = None
            row[prefix + "current_rx_mode"] = None

    # Parse warnings
    warnings = event.get("parse_warnings", [])
    row["parse_warnings"] = "; ".join(warnings) if warnings else None

    # Sanitize all string values to remove illegal XML characters
    return {k: sanitize(v) for k, v in row.items()}


def main():
    jsonl_files = sorted(glob.glob(os.path.join(EVENTS_DIR, "*.jsonl")))
    print(f"Found {len(jsonl_files)} JSONL files")

    all_rows = []
    for filepath in jsonl_files:
        year_file = os.path.basename(filepath)
        count = 0
        with open(filepath, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                event = json.loads(line)
                all_rows.append(flatten_event(event))
                count += 1
        print(f"  {year_file}: {count} events")

    print(f"\nTotal events: {len(all_rows)}")

    # Define column order
    column_order = [
        "event_number", "event_year", "category", "emergency_class",
        "report_date", "event_date", "event_time", "event_timezone",
        "notification_date", "notification_time", "notification_timezone",
        "last_update_date",
        "facility", "region", "state", "county", "city",
        "unit", "rx_type",
        "licensee", "license_number", "rep_org", "agreement", "docket",
        "cfr_codes", "cfr_descriptions",
        "persons_notified_names", "persons_notified_orgs",
        "reactor_unit_1_unit", "reactor_unit_1_scram_code", "reactor_unit_1_rx_crit",
        "reactor_unit_1_initial_power", "reactor_unit_1_initial_rx_mode",
        "reactor_unit_1_current_power", "reactor_unit_1_current_rx_mode",
        "reactor_unit_2_unit", "reactor_unit_2_scram_code", "reactor_unit_2_rx_crit",
        "reactor_unit_2_initial_power", "reactor_unit_2_initial_rx_mode",
        "reactor_unit_2_current_power", "reactor_unit_2_current_rx_mode",
        "reactor_unit_3_unit", "reactor_unit_3_scram_code", "reactor_unit_3_rx_crit",
        "reactor_unit_3_initial_power", "reactor_unit_3_initial_rx_mode",
        "reactor_unit_3_current_power", "reactor_unit_3_current_rx_mode",
        "event_text",
        "page_url", "scraped_at", "html_format", "parse_warnings",
    ]

    df = pd.DataFrame(all_rows, columns=column_order)
    df.sort_values(by=["event_number"], ascending=True, inplace=True)
    df.reset_index(drop=True, inplace=True)

    # Write to Excel with pandas first (for data), then format with openpyxl
    print(f"Writing {len(df)} rows to {OUTPUT_FILE}...")
    df.to_excel(OUTPUT_FILE, index=False, sheet_name="NRC Events", engine="openpyxl")

    # Format with openpyxl
    wb = load_workbook(OUTPUT_FILE)
    ws = wb["NRC Events"]

    # Header styling (only header row — skip per-cell formatting for speed)
    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Set reasonable column widths
    width_map = {
        "event_number": 14, "event_year": 10, "category": 18, "emergency_class": 18,
        "report_date": 12, "event_date": 12, "event_time": 10, "event_timezone": 10,
        "notification_date": 14, "notification_time": 14, "notification_timezone": 14,
        "last_update_date": 16,
        "facility": 22, "region": 8, "state": 8, "county": 14, "city": 14,
        "unit": 14, "rx_type": 22,
        "licensee": 30, "license_number": 16, "rep_org": 30, "agreement": 10, "docket": 12,
        "cfr_codes": 30, "cfr_descriptions": 40,
        "persons_notified_names": 25, "persons_notified_orgs": 20,
        "event_text": 60, "page_url": 35, "scraped_at": 22, "html_format": 12,
        "parse_warnings": 20,
    }
    # Reactor unit columns
    for i in range(1, 4):
        prefix = f"reactor_unit_{i}_"
        width_map[prefix + "unit"] = 8
        width_map[prefix + "scram_code"] = 12
        width_map[prefix + "rx_crit"] = 10
        width_map[prefix + "initial_power"] = 14
        width_map[prefix + "initial_rx_mode"] = 18
        width_map[prefix + "current_power"] = 14
        width_map[prefix + "current_rx_mode"] = 18

    for col_idx, col_name in enumerate(column_order, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width_map.get(col_name, 14)

    # Freeze top row and first 2 columns (event_number, event_year)
    ws.freeze_panes = "C2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT_FILE)
    print(f"Done! File saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
